from copy import copy
from datetime import date, datetime
from hashlib import sha1
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.auth.security import get_current_user, require_roles
from app.core.config import settings
from app.core.database import get_db
from app.models.entities import Mapping, Role, User
from app.services.audit import write_audit
from app.services.mappings import create_backup

router = APIRouter(tags=["Import / Export"])
PRIMARY_SHEET = "Blueprint_โพธิ์ไทร"
HEADERS = [
    "namepttype",
    "rights",
    "pttype",
    "stdcode",
    "pay",
    "op56",
    "inscl",
    "chkshow",
    "instypeold",
    "rightgroup",
    "chg18",
    "income_id",
    "que_group",
    "sss_export",
    "j_export",
    "ins_code",
    "dept_code_id",
    "dept_cr_code_id",
    "dept_code_ip_id",
    "dept_rf_code_id",
    "dept_rfo_code_id",
    "pttype_replace",
    "รหัสผังบัญชีลูกหนี้_OP",
    "รหัสผังบัญชีรายได้_OP",
    "รหัสผังบัญชีลูกหนี้_IP",
    "รหัสผังบัญชีรายได้_IP",
    "ชื่อบัญชี OP",
    "ชื่อบัญชี IP",
    "หมายเหตุ",
]


def json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def canonical_values(data: dict[str, Any]) -> dict[str, str]:
    benefit_code = str(data.get("pttype") or "").strip()
    benefit_name = str(data.get("namepttype") or "").strip()
    if not benefit_code:
        identity = "|".join(
            str(data.get(field) or "")
            for field in ("namepttype", "inscl", "รหัสผังบัญชีลูกหนี้_OP", "รหัสผังบัญชีลูกหนี้_IP")
        )
        benefit_code = f"__BLANK_{sha1(identity.encode('utf-8')).hexdigest()[:12]}"
    account_code = str(
        data.get("รหัสผังบัญชีลูกหนี้_OP")
        or data.get("รหัสผังบัญชีลูกหนี้_IP")
        or f"UNMAPPED:{benefit_code}"
    ).strip()
    account_name = str(
        data.get("ชื่อบัญชี OP") or data.get("ชื่อบัญชี IP") or benefit_name
    ).strip()
    return {
        "benefit_code": benefit_code,
        "benefit_name": benefit_name,
        "account_code": account_code,
        "account_name": account_name,
    }


@router.post("/import")
async def import_excel(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(Role.editor, Role.approver, Role.super_admin)),
) -> dict:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    content = await file.read(settings.max_upload_mb * 1024 * 1024 + 1)
    if len(content) > settings.max_upload_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Uploaded file is too large")
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, read_only=True)
        sheet = workbook[PRIMARY_SHEET] if PRIMARY_SHEET in workbook.sheetnames else workbook.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid Excel file") from exc
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows())]
    missing = [header for header in HEADERS if header not in headers]
    if missing:
        raise HTTPException(status_code=422, detail={"missing_headers": missing, "required_headers": HEADERS})
    index = {name: headers.index(name) for name in headers}
    parsed_rows: list[dict[str, Any]] = []
    errors: list[dict] = []
    seen_pttypes: set[str] = set()
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None for value in cells):
            continue
        data = {header: json_value(cells[index[header]]) for header in HEADERS}
        canonical = canonical_values(data)
        if canonical["benefit_code"] in seen_pttypes:
            errors.append({"row": row_number, "error": f"duplicate pttype: {canonical['benefit_code']}"})
        if not canonical["benefit_name"]:
            errors.append({"row": row_number, "error": "namepttype is required"})
        seen_pttypes.add(canonical["benefit_code"])
        parsed_rows.append({"row_number": row_number, "data": data, **canonical})
    if errors:
        raise HTTPException(status_code=422, detail={"message": "Import aborted; no rows were saved", "errors": errors[:100]})

    create_backup(db, user, f"before-import:{file.filename}")
    existing = {
        item.benefit_code: item
        for item in db.scalars(select(Mapping).where(Mapping.is_deleted.is_(False))).all()
    }
    imported = 0
    updated = 0
    imported_codes: set[str] = set()
    for sequence, row in enumerate(parsed_rows, start=1):
        item = existing.get(row["benefit_code"])
        if item:
            updated += 1
        else:
            imported += 1
            item = Mapping(
                created_by=user.id,
                updated_by=user.id,
                effective_date=date.today(),
                benefit_code=row["benefit_code"],
                benefit_name=row["benefit_name"],
                account_code=row["account_code"],
                account_name=row["account_name"],
            )
            db.add(item)
        item.sequence = sequence
        item.benefit_code = row["benefit_code"]
        item.benefit_name = row["benefit_name"]
        item.account_code = row["account_code"]
        item.account_name = row["account_name"]
        item.description = str(row["data"].get("หมายเหตุ") or "") or None
        item.source_sheet = sheet.title
        item.source_row = row["row_number"]
        item.source_data = row["data"]
        item.updated_by = user.id
        if item.id:
            item.version += 1
        imported_codes.add(row["benefit_code"])
    deleted = 0
    for code, item in existing.items():
        if code not in imported_codes:
            item.is_deleted = True
            item.updated_by = user.id
            item.version += 1
            deleted += 1
    write_audit(
        db,
        request,
        user,
        "import",
        "mapping",
        new_value={
            "filename": file.filename,
            "sheet": sheet.title,
            "inserted": imported,
            "updated": updated,
            "deleted": deleted,
            "record_count": len(parsed_rows),
        },
    )
    db.commit()
    return {
        "filename": file.filename,
        "sheet": sheet.title,
        "inserted": imported,
        "updated": updated,
        "deleted": deleted,
        "record_count": len(parsed_rows),
    }


@router.get("/export")
def export_excel(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
) -> StreamingResponse:
    items = db.scalars(
        select(Mapping).where(Mapping.is_deleted.is_(False)).order_by(Mapping.sequence)
    ).all()
    template_path = Path(settings.excel_template_path)
    if not template_path.exists():
        raise HTTPException(status_code=500, detail="Excel template is not installed")
    workbook = load_workbook(template_path)
    sheet = workbook[PRIMARY_SHEET]
    style_source = [copy(sheet.cell(2, column)._style) for column in range(1, len(HEADERS) + 1)]
    alignments = [copy(sheet.cell(2, column).alignment) for column in range(1, len(HEADERS) + 1)]
    fills = [copy(sheet.cell(2, column).fill) for column in range(1, len(HEADERS) + 1)]
    fonts = [copy(sheet.cell(2, column).font) for column in range(1, len(HEADERS) + 1)]
    borders = [copy(sheet.cell(2, column).border) for column in range(1, len(HEADERS) + 1)]
    number_formats = [sheet.cell(2, column).number_format for column in range(1, len(HEADERS) + 1)]
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for row_number, item in enumerate(items, start=2):
        data = item.source_data or {}
        data.setdefault("namepttype", item.benefit_name)
        data.setdefault("pttype", item.benefit_code)
        data.setdefault("รหัสผังบัญชีลูกหนี้_OP", item.account_code)
        data.setdefault("ชื่อบัญชี OP", item.account_name)
        data.setdefault("หมายเหตุ", item.description)
        for column, header in enumerate(HEADERS, start=1):
            cell = sheet.cell(row_number, column, data.get(header))
            cell._style = copy(style_source[column - 1])
            cell.alignment = copy(alignments[column - 1])
            cell.fill = copy(fills[column - 1])
            cell.font = copy(fonts[column - 1])
            cell.border = copy(borders[column - 1])
            cell.number_format = number_formats[column - 1]
        inscl = str(data.get("inscl") or "")
        note = str(data.get("หมายเหตุ") or "")
        sheet.row_dimensions[row_number].hidden = inscl not in {"BKK", "NRH", "LGO", "BMT", "OFC"} or note != "มีใช้"
    last_row = max(2, len(items) + 1)
    sheet.auto_filter.ref = f"A1:AC{last_row}"
    sheet.data_validations.dataValidation = []
    validation = DataValidation(type="list", formula1='"มีใช้,ไม่ใช้แล้ว"', allow_blank=True)
    validation.add(f"AC2:AC{last_row}")
    sheet.add_data_validation(validation)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ub-amms-mappings.xlsx"'},
    )
